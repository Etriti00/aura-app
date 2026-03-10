"""
Aura v2.0 — Professional Excel Export Engine
openpyxl-based .xlsx exports with formulas, charts, styled headers.
"""

from pathlib import Path
from datetime import datetime

from utils.logger import get_logger

logger = get_logger(__name__)

# Lazy import openpyxl — it's optional
_openpyxl = None


def _get_openpyxl():
    global _openpyxl
    if _openpyxl is None:
        try:
            import openpyxl
            _openpyxl = openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel exports: pip install openpyxl")
    return _openpyxl


# ─── Shared Styling ──────────────────────────────────────────────────────────

_HEADER_COLOR = "1F4E79"
_HEADER_FONT_COLOR = "FFFFFF"
_ALT_ROW_COLOR = "F2F7FB"
_CURRENCY_FMT = '"$"#,##0.00'
_PCT_FMT = '0.0%'


def _style_header_row(ws, row_num: int, num_cols: int):
    """Apply dark blue header styling to a row."""
    openpyxl = _get_openpyxl()
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill(start_color=_HEADER_COLOR, end_color=_HEADER_COLOR, fill_type="solid")
    header_font = Font(color=_HEADER_FONT_COLOR, bold=True, size=11)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)


def _alt_row_fill(ws, start_row: int, end_row: int, num_cols: int):
    """Apply alternating row colors."""
    openpyxl = _get_openpyxl()
    from openpyxl.styles import PatternFill

    alt_fill = PatternFill(start_color=_ALT_ROW_COLOR, end_color=_ALT_ROW_COLOR, fill_type="solid")
    for row in range(start_row, end_row + 1):
        if row % 2 == 0:
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).fill = alt_fill


class ExcelExportEngine:
    """Professional Excel export engine for campaigns, leads, pipeline, finance."""

    def __init__(self, db_manager):
        self.db = db_manager

    def export_campaign_xlsx(self, campaign_id: int, output_path: str) -> str:
        """Export a campaign to a 4-sheet Excel workbook.

        Sheets: Summary, Leads, Outreach, Cost Breakdown
        """
        openpyxl = _get_openpyxl()
        from openpyxl.styles import Font, Alignment, numbers
        from database.schema import Campaign, Lead, EnrichmentData

        wb = openpyxl.Workbook()

        with self.db.session_scope() as session:
            campaign = session.query(Campaign).filter_by(id=campaign_id).first()
            if not campaign:
                raise ValueError(f"Campaign {campaign_id} not found")

            leads = session.query(Lead).filter_by(campaign_id=campaign_id).all()

            # ─── Sheet 1: Summary ─────────────────────────
            ws_summary = wb.active
            ws_summary.title = "Summary"

            ws_summary.cell(row=1, column=1, value="Campaign Report")
            ws_summary.cell(row=1, column=1).font = Font(size=16, bold=True, color=_HEADER_COLOR)
            ws_summary.cell(row=2, column=1, value=campaign.name)
            ws_summary.cell(row=2, column=1).font = Font(size=12)
            ws_summary.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

            # KPIs
            kpis = [
                ("Total Leads", campaign.total_leads),
                ("Qualified Leads", campaign.qualified_leads),
                ("Emails Sent", campaign.emails_sent),
                ("Replies", campaign.replies),
                ("Reply Rate", campaign.replies / max(campaign.emails_sent, 1)),
                ("Total Cost", campaign.total_token_cost),
                ("Cost per Lead", campaign.total_token_cost / max(campaign.total_leads, 1)),
            ]
            headers = ["Metric", "Value"]
            ws_summary.append([])
            ws_summary.append(headers)
            _style_header_row(ws_summary, ws_summary.max_row, 2)
            for name, value in kpis:
                ws_summary.append([name, value])
                if "Rate" in name:
                    ws_summary.cell(row=ws_summary.max_row, column=2).number_format = _PCT_FMT
                elif "Cost" in name:
                    ws_summary.cell(row=ws_summary.max_row, column=2).number_format = _CURRENCY_FMT

            _auto_width(ws_summary)

            # ─── Sheet 2: Leads ───────────────────────────
            ws_leads = wb.create_sheet("Leads")
            lead_headers = ["ID", "Business", "Category", "City", "Email", "Phone",
                            "Status", "Website Score", "Lifecycle"]
            ws_leads.append(lead_headers)
            _style_header_row(ws_leads, 1, len(lead_headers))
            ws_leads.freeze_panes = "A2"
            ws_leads.auto_filter.ref = f"A1:{chr(64 + len(lead_headers))}1"

            for lead in leads:
                ws_leads.append([
                    lead.id, lead.business_name, lead.category, lead.city,
                    lead.email, lead.phone, lead.status, lead.website_score,
                    lead.lifecycle_state,
                ])

            _alt_row_fill(ws_leads, 2, ws_leads.max_row, len(lead_headers))
            _auto_width(ws_leads)

            # ─── Sheet 3: Outreach ────────────────────────
            ws_outreach = wb.create_sheet("Outreach")
            outreach_headers = ["Business", "Email", "Subject", "Sent At", "Status"]
            ws_outreach.append(outreach_headers)
            _style_header_row(ws_outreach, 1, len(outreach_headers))

            sent_leads = [l for l in leads if l.email_sent_at]
            for lead in sent_leads:
                ws_outreach.append([
                    lead.business_name, lead.email,
                    lead.email_subject[:80] if lead.email_subject else "",
                    lead.email_sent_at.strftime("%Y-%m-%d %H:%M") if lead.email_sent_at else "",
                    lead.status,
                ])

            _alt_row_fill(ws_outreach, 2, ws_outreach.max_row, len(outreach_headers))
            _auto_width(ws_outreach)

            # ─── Sheet 4: Cost Breakdown ──────────────────
            ws_cost = wb.create_sheet("Cost Breakdown")
            cost_headers = ["Tier", "Leads", "Cost USD"]
            ws_cost.append(cost_headers)
            _style_header_row(ws_cost, 1, len(cost_headers))

            tier_costs = {}
            for lead in leads:
                if lead.tier2_cost_usd > 0:
                    tier_costs.setdefault("Tier 2 (Flash)", [0, 0.0])
                    tier_costs["Tier 2 (Flash)"][0] += 1
                    tier_costs["Tier 2 (Flash)"][1] += lead.tier2_cost_usd
                if lead.tier3_cost_usd > 0:
                    tier_costs.setdefault("Tier 3 (Sonnet)", [0, 0.0])
                    tier_costs["Tier 3 (Sonnet)"][0] += 1
                    tier_costs["Tier 3 (Sonnet)"][1] += lead.tier3_cost_usd

            for tier, (count, cost) in tier_costs.items():
                ws_cost.append([tier, count, cost])
                ws_cost.cell(row=ws_cost.max_row, column=3).number_format = _CURRENCY_FMT

            # Total row
            total_row = ws_cost.max_row + 1
            ws_cost.cell(row=total_row, column=1, value="TOTAL")
            ws_cost.cell(row=total_row, column=1).font = Font(bold=True)
            ws_cost.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})")
            ws_cost.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row - 1})")
            ws_cost.cell(row=total_row, column=3).number_format = _CURRENCY_FMT

            _auto_width(ws_cost)

            # ─── Chart: Cost per Tier ───────────────────────
            if tier_costs:
                try:
                    from openpyxl.chart import BarChart, Reference
                    chart = BarChart()
                    chart.title = "Cost per Tier"
                    chart.y_axis.title = "USD"
                    chart.x_axis.title = "Tier"
                    chart.style = 10
                    data_rows = len(tier_costs)
                    data_ref = Reference(ws_cost, min_col=3, min_row=1, max_row=data_rows + 1)
                    cats_ref = Reference(ws_cost, min_col=1, min_row=2, max_row=data_rows + 1)
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cats_ref)
                    chart.width = 20
                    chart.height = 12
                    ws_cost.add_chart(chart, "E2")
                except Exception as e:
                    logger.debug(f"Chart creation failed: {e}")

        wb.save(output_path)
        logger.debug(f"Campaign XLSX exported to {output_path}")
        return output_path

    def export_pipeline_xlsx(self, output_path: str) -> str:
        """Export cross-campaign pipeline funnel with conversion rate formulas."""
        openpyxl = _get_openpyxl()
        from openpyxl.styles import Font
        from database.schema import Campaign

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pipeline"

        headers = ["Campaign", "Leads", "Qualified", "Emailed", "Replies",
                    "Qual Rate", "Reply Rate"]
        ws.append(headers)
        _style_header_row(ws, 1, len(headers))

        with self.db.session_scope() as session:
            campaigns = session.query(Campaign).order_by(Campaign.created_at.desc()).all()
            for i, c in enumerate(campaigns):
                row_num = i + 2
                ws.append([
                    c.name, c.total_leads, c.qualified_leads, c.emails_sent, c.replies,
                    None, None,
                ])
                # Qual Rate formula: =C/B
                ws.cell(row=row_num, column=6, value=f"=IF(B{row_num}>0,C{row_num}/B{row_num},0)")
                ws.cell(row=row_num, column=6).number_format = _PCT_FMT
                # Reply Rate formula: =E/D
                ws.cell(row=row_num, column=7, value=f"=IF(D{row_num}>0,E{row_num}/D{row_num},0)")
                ws.cell(row=row_num, column=7).number_format = _PCT_FMT

        # Totals
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        for col in range(2, 6):
            ws.cell(row=total_row, column=col, value=f"=SUM({chr(64+col)}2:{chr(64+col)}{total_row-1})")

        _alt_row_fill(ws, 2, ws.max_row, len(headers))
        _auto_width(ws)

        wb.save(output_path)
        logger.debug(f"Pipeline XLSX exported to {output_path}")
        return output_path

    def export_lead_detail_xlsx(self, lead_id: int, output_path: str) -> str:
        """Export detailed lead data with enrichment + lifecycle."""
        openpyxl = _get_openpyxl()
        from openpyxl.styles import Font
        from database.schema import Lead, EnrichmentData, LeadStateTransition, CaseNote

        wb = openpyxl.Workbook()

        with self.db.session_scope() as session:
            lead = session.query(Lead).filter_by(id=lead_id).first()
            if not lead:
                raise ValueError(f"Lead {lead_id} not found")

            # ─── Sheet 1: Lead Info ──────────────────────
            ws = wb.active
            ws.title = "Lead Info"

            fields = [
                ("Business Name", lead.business_name),
                ("Category", lead.category),
                ("City", lead.city),
                ("Country", lead.country),
                ("Email", lead.email),
                ("Phone", lead.phone),
                ("Website", lead.website_url),
                ("Website Score", lead.website_score),
                ("Status", lead.status),
                ("Lifecycle State", lead.lifecycle_state),
                ("Completeness Score", lead.data_completeness_score),
                ("Created", lead.created_at.strftime("%Y-%m-%d") if lead.created_at else ""),
            ]

            ws.append(["Field", "Value"])
            _style_header_row(ws, 1, 2)
            for name, value in fields:
                ws.append([name, value])
            _auto_width(ws)

            # ─── Sheet 2: Enrichment ─────────────────────
            enrichment = session.query(EnrichmentData).filter_by(lead_id=lead_id).first()
            if enrichment:
                ws_e = wb.create_sheet("Enrichment")
                ws_e.append(["Field", "Value"])
                _style_header_row(ws_e, 1, 2)
                enrichment_fields = [
                    ("Google Maps Rating", enrichment.google_maps_rating),
                    ("Domain Age (years)", enrichment.domain_age_years),
                    ("Has SSL", enrichment.has_ssl),
                    ("MX Records Valid", enrichment.mx_records_valid),
                    ("Tech Stack", enrichment.tech_stack),
                    ("Decision Maker", enrichment.decision_maker_name),
                    ("Decision Maker Title", enrichment.decision_maker_title),
                    ("Company Description", enrichment.company_description),
                    ("ICP Fit Score", enrichment.icp_fit_score),
                    ("Industry", enrichment.industry_tag),
                    ("Company Size", enrichment.company_size_estimate),
                ]
                for name, value in enrichment_fields:
                    ws_e.append([name, value])
                _auto_width(ws_e)

            # ─── Sheet 3: Lifecycle ──────────────────────
            transitions = (
                session.query(LeadStateTransition)
                .filter_by(lead_id=lead_id)
                .order_by(LeadStateTransition.created_at)
                .all()
            )
            if transitions:
                ws_l = wb.create_sheet("Lifecycle")
                ws_l.append(["From", "To", "Triggered By", "Date"])
                _style_header_row(ws_l, 1, 4)
                for t in transitions:
                    ws_l.append([
                        t.from_state, t.to_state, t.triggered_by,
                        t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
                    ])
                _auto_width(ws_l)

            # ─── Sheet 4: Case Notes ─────────────────────
            notes = (
                session.query(CaseNote)
                .filter_by(lead_id=lead_id)
                .order_by(CaseNote.created_at.desc())
                .limit(50)
                .all()
            )
            if notes:
                ws_n = wb.create_sheet("Case Notes")
                ws_n.append(["Type", "Content", "Date"])
                _style_header_row(ws_n, 1, 3)
                for n in notes:
                    ws_n.append([
                        n.note_type, n.content[:500],
                        n.created_at.strftime("%Y-%m-%d %H:%M") if n.created_at else "",
                    ])
                _auto_width(ws_n)

        wb.save(output_path)
        logger.debug(f"Lead detail XLSX exported to {output_path}")
        return output_path

    def export_finance_xlsx(self, output_path: str) -> str:
        """Export finance data: invoices, revenue, clients."""
        openpyxl = _get_openpyxl()
        from openpyxl.styles import Font
        from database.schema import Invoice, InvoiceLineItem

        wb = openpyxl.Workbook()

        with self.db.session_scope() as session:
            # ─── Sheet 1: Invoices ──────────────────────
            ws = wb.active
            ws.title = "Invoices"
            headers = ["Invoice #", "Client", "Email", "Subtotal", "Tax", "Total",
                        "Currency", "Status", "Due Date", "Created"]
            ws.append(headers)
            _style_header_row(ws, 1, len(headers))

            invoices = session.query(Invoice).order_by(Invoice.created_at.desc()).all()
            for inv in invoices:
                ws.append([
                    inv.invoice_number, inv.client_name, inv.client_email,
                    inv.subtotal, inv.tax_amount, inv.total,
                    inv.currency, inv.status,
                    inv.due_date.strftime("%Y-%m-%d") if inv.due_date else "",
                    inv.created_at.strftime("%Y-%m-%d") if inv.created_at else "",
                ])
                for col in [4, 5, 6]:
                    ws.cell(row=ws.max_row, column=col).number_format = _CURRENCY_FMT

            # Total row
            if invoices:
                total_row = ws.max_row + 1
                ws.cell(row=total_row, column=1, value="TOTAL")
                ws.cell(row=total_row, column=1).font = Font(bold=True)
                ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row-1})")
                ws.cell(row=total_row, column=6).number_format = _CURRENCY_FMT

            _alt_row_fill(ws, 2, ws.max_row, len(headers))
            _auto_width(ws)

            # ─── Sheet 2: Revenue Summary ────────────────
            ws_rev = wb.create_sheet("Revenue")
            ws_rev.append(["Status", "Count", "Total"])
            _style_header_row(ws_rev, 1, 3)

            status_totals = {}
            for inv in invoices:
                status_totals.setdefault(inv.status, [0, 0.0])
                status_totals[inv.status][0] += 1
                status_totals[inv.status][1] += inv.total

            for status, (count, total) in status_totals.items():
                ws_rev.append([status, count, total])
                ws_rev.cell(row=ws_rev.max_row, column=3).number_format = _CURRENCY_FMT

            _auto_width(ws_rev)

        wb.save(output_path)
        logger.debug(f"Finance XLSX exported to {output_path}")
        return output_path
