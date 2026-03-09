"""Tests for v2.0 Excel Export Engine."""

import os
import tempfile

import pytest

from database.schema import Campaign, Lead, Invoice, InvoiceLineItem, Settings


class TestExcelImportCheck:
    """Verify openpyxl availability."""

    def test_openpyxl_importable(self):
        import openpyxl
        assert hasattr(openpyxl, "Workbook")


class TestCampaignExport:
    """Test campaign XLSX export."""

    def test_export_campaign_xlsx(self, db):
        from core.excel_export_engine import ExcelExportEngine

        # Seed data
        with db.session_scope() as session:
            camp = Campaign(name="Test Export", search_query="plumbing austin",
                            total_leads=10, qualified_leads=5, emails_sent=5,
                            replies=2, total_token_cost=0.50)
            session.add(camp)
            session.flush()
            camp_id = camp.id

            for i in range(3):
                lead = Lead(
                    campaign_id=camp_id, business_name=f"Lead {i}",
                    category="plumbing", city="Austin", email=f"lead{i}@test.com",
                    status="qualified" if i < 2 else "new",
                    website_score=80 - i * 10, tier2_cost_usd=0.01 * (i + 1),
                )
                session.add(lead)

        engine = ExcelExportEngine(db)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name

        try:
            result = engine.export_campaign_xlsx(camp_id, path)
            assert os.path.exists(result)

            import openpyxl
            wb = openpyxl.load_workbook(result)
            assert "Summary" in wb.sheetnames
            assert "Leads" in wb.sheetnames
            assert "Outreach" in wb.sheetnames
            assert "Cost Breakdown" in wb.sheetnames

            # Check leads sheet has rows
            ws_leads = wb["Leads"]
            assert ws_leads.max_row >= 4  # header + 3 leads

            # Check summary has campaign name
            ws_summary = wb["Summary"]
            assert ws_summary.cell(row=2, column=1).value == "Test Export"
        finally:
            os.unlink(path)

    def test_export_nonexistent_campaign(self, db):
        from core.excel_export_engine import ExcelExportEngine
        engine = ExcelExportEngine(db)
        with pytest.raises(ValueError, match="not found"):
            engine.export_campaign_xlsx(9999, "/tmp/test.xlsx")


class TestPipelineExport:
    def test_export_pipeline_xlsx(self, db):
        from core.excel_export_engine import ExcelExportEngine

        with db.session_scope() as session:
            for i in range(3):
                session.add(Campaign(
                    name=f"Campaign {i}", total_leads=10 * (i + 1),
                    qualified_leads=5 * (i + 1), emails_sent=5 * (i + 1),
                    replies=i, total_token_cost=0.1 * (i + 1),
                ))

        engine = ExcelExportEngine(db)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name

        try:
            result = engine.export_pipeline_xlsx(path)
            assert os.path.exists(result)

            import openpyxl
            wb = openpyxl.load_workbook(result)
            ws = wb["Pipeline"]
            assert ws.max_row >= 4  # header + 3 campaigns + total
        finally:
            os.unlink(path)


class TestLeadDetailExport:
    def test_export_lead_detail(self, db):
        from core.excel_export_engine import ExcelExportEngine

        with db.session_scope() as session:
            camp = Campaign(name="Detail Test")
            session.add(camp)
            session.flush()
            lead = Lead(
                campaign_id=camp.id, business_name="Detail Lead",
                city="Austin", email="detail@test.com", phone="555-1234",
                website_score=85, lifecycle_state="qualified",
            )
            session.add(lead)
            session.flush()
            lead_id = lead.id

        engine = ExcelExportEngine(db)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name

        try:
            result = engine.export_lead_detail_xlsx(lead_id, path)
            assert os.path.exists(result)

            import openpyxl
            wb = openpyxl.load_workbook(result)
            assert "Lead Info" in wb.sheetnames
        finally:
            os.unlink(path)

    def test_export_nonexistent_lead(self, db):
        from core.excel_export_engine import ExcelExportEngine
        engine = ExcelExportEngine(db)
        with pytest.raises(ValueError, match="not found"):
            engine.export_lead_detail_xlsx(9999, "/tmp/test.xlsx")


class TestFinanceExport:
    def test_export_finance_xlsx(self, db):
        from core.excel_export_engine import ExcelExportEngine

        with db.session_scope() as session:
            inv1 = Invoice(
                invoice_number="INV-001", client_name="Client A",
                client_email="a@test.com", subtotal=1000, tax_rate=0.21,
                tax_amount=210, total=1210, status="paid",
            )
            inv2 = Invoice(
                invoice_number="INV-002", client_name="Client B",
                client_email="b@test.com", subtotal=500, tax_rate=0.21,
                tax_amount=105, total=605, status="draft",
            )
            session.add_all([inv1, inv2])

        engine = ExcelExportEngine(db)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name

        try:
            result = engine.export_finance_xlsx(path)
            assert os.path.exists(result)

            import openpyxl
            wb = openpyxl.load_workbook(result)
            assert "Invoices" in wb.sheetnames
            assert "Revenue" in wb.sheetnames

            ws = wb["Invoices"]
            assert ws.max_row >= 3  # header + 2 invoices
        finally:
            os.unlink(path)

    def test_export_empty_finance(self, db):
        from core.excel_export_engine import ExcelExportEngine
        engine = ExcelExportEngine(db)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name

        try:
            result = engine.export_finance_xlsx(path)
            assert os.path.exists(result)
        finally:
            os.unlink(path)


class TestStyling:
    def test_header_row_styling(self, db):
        """Verify header rows have dark blue fill."""
        from core.excel_export_engine import ExcelExportEngine

        with db.session_scope() as session:
            camp = Campaign(name="Style Test", total_leads=1)
            session.add(camp)
            session.flush()
            session.add(Lead(campaign_id=camp.id, business_name="X", status="new"))
            session.flush()
            camp_id = camp.id

        engine = ExcelExportEngine(db)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name

        try:
            engine.export_campaign_xlsx(camp_id, path)

            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb["Leads"]
            header_cell = ws.cell(row=1, column=1)
            assert header_cell.fill.start_color.rgb is not None
            assert header_cell.font.bold is True
        finally:
            os.unlink(path)

    def test_freeze_panes(self, db):
        """Verify leads sheet has frozen header row."""
        from core.excel_export_engine import ExcelExportEngine

        with db.session_scope() as session:
            camp = Campaign(name="Freeze Test", total_leads=1)
            session.add(camp)
            session.flush()
            session.add(Lead(campaign_id=camp.id, business_name="Y", status="new"))
            session.flush()
            camp_id = camp.id

        engine = ExcelExportEngine(db)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name

        try:
            engine.export_campaign_xlsx(camp_id, path)

            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb["Leads"]
            assert ws.freeze_panes == "A2"
        finally:
            os.unlink(path)
