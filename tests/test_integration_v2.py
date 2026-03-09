"""
Integration tests for v2.0 features — end-to-end flows across multiple engines.
Tests enrichment pipeline, invoice lifecycle, response formatting, and exports.
"""

import os
import tempfile
from datetime import datetime, timedelta
from unittest.mock import MagicMock

import pytest

from database.schema import (
    Campaign, Lead, EnrichmentData, Service, Invoice, Settings,
)


class TestEnrichmentPipeline:
    """Test the 5-layer enrichment scoring flow."""

    def test_completeness_scoring(self, db):
        from core.enrichment_scoring import calculate_completeness

        # Empty enrichment
        assert calculate_completeness({}) == 0

        # Email only (40 weight)
        assert calculate_completeness({"email": "test@test.com"}) == 40

        # Full enrichment
        full = {
            "email": "test@test.com",
            "decision_maker_name": "John Doe",
            "pain_points": "Website is slow",
            "tech_stack": "WordPress",
            "phone": "555-1234",
        }
        assert calculate_completeness(full) == 100

    def test_layer0_dns_basic(self):
        from core.enrichment_layers.layer0_dns import enrich_layer0

        html = """
        <html>
        <head>
            <meta name="generator" content="WordPress 6.0">
            <meta name="viewport" content="width=device-width">
        </head>
        <body>
            <a href="https://linkedin.com/company/test">LinkedIn</a>
            <a href="https://twitter.com/test">Twitter</a>
        </body>
        </html>
        """
        result = enrich_layer0("example.com", html)
        assert "WordPress" in result.get("tech_stack", "")
        assert result.get("is_mobile_responsive") is True
        social = result.get("social_links", "")
        assert "linkedin" in social.lower()

    def test_scoring_after_partial_enrichment(self, db):
        from core.enrichment_scoring import calculate_completeness

        # Simulates Layer 0 + 2 results (no email yet)
        partial = {
            "tech_stack": "WordPress, jQuery",
            "social_links": "linkedin.com/company/test",
        }
        score = calculate_completeness(partial)
        assert score == 10  # Only tech_stack (10 weight)

        # After Layer 3 adds email
        partial["email"] = "test@test.com"
        score = calculate_completeness(partial)
        assert score == 50  # email(40) + tech(10)


class TestInvoiceLifecycle:
    """Test the full invoice flow: create → approve → PDF → paid."""

    def test_full_invoice_flow(self, db):
        from core.pricing_engine import PricingEngine
        from core.invoice_approval_engine import InvoiceApprovalEngine

        pe = PricingEngine(db)
        iae = InvoiceApprovalEngine(db, pricing_engine=pe)

        # Seed a service
        with db.session_scope() as session:
            session.add(Service(
                name="Web Design", base_price=1000, max_price=5000,
                category="development",
            ))

        # Create invoice with line items
        result = pe.create_invoice(
            client_name="Full Flow Corp",
            client_email="full@corp.com",
            line_items=[
                {"description": "Web Design", "quantity": 1, "unit_price": 2000},
                {"description": "SEO Setup", "quantity": 1, "unit_price": 500},
            ],
        )
        assert result["success"]
        inv_id = result["data"]["id"]
        assert result["data"]["subtotal"] == 2500.0

        # Start approval
        approval = iae.start_approval_flow(inv_id)
        assert approval["success"]

        # Approve
        approve = iae.approve(inv_id, approved_by="admin")
        assert approve["success"]
        assert approve["data"]["approval_status"] == "approved"

        # Generate PDF
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            pdf_path = f.name

        try:
            pdf = pe.generate_invoice_pdf(inv_id, pdf_path)
            assert pdf["success"]
            assert os.path.exists(pdf_path)
        finally:
            os.unlink(pdf_path)

        # Mark as paid
        paid = pe.update_status(inv_id, "paid")
        assert paid["success"]

        # Verify final state
        invoice = pe.get_invoice(inv_id)
        assert invoice["data"]["status"] == "paid"
        assert invoice["data"]["approval_status"] == "approved"

    def test_rejected_invoice_flow(self, db):
        from core.pricing_engine import PricingEngine
        from core.invoice_approval_engine import InvoiceApprovalEngine

        pe = PricingEngine(db)
        iae = InvoiceApprovalEngine(db)

        result = pe.create_invoice(
            client_name="Rejected Corp",
            line_items=[{"description": "X", "quantity": 1, "unit_price": 10000}],
        )
        inv_id = result["data"]["id"]

        reject = iae.reject(inv_id, reason="Price too high")
        assert reject["success"]

        invoice = pe.get_invoice(inv_id)
        assert invoice["data"]["approval_status"] == "rejected"

    def test_revenue_after_invoices(self, db):
        from core.pricing_engine import PricingEngine

        pe = PricingEngine(db)

        # Create and pay 2 invoices
        inv1 = pe.create_invoice(
            client_name="A",
            line_items=[{"description": "X", "quantity": 1, "unit_price": 1000}],
        )
        inv2 = pe.create_invoice(
            client_name="B",
            line_items=[{"description": "Y", "quantity": 1, "unit_price": 2000}],
        )
        pe.update_status(inv1["data"]["id"], "paid")

        rev = pe.get_revenue_summary()
        assert rev["data"]["paid_count"] == 1
        assert rev["data"]["total_revenue"] == 1000.0
        assert rev["data"]["total_pending"] == 2000.0


class TestResponseFormatting:
    """Test rendering across all 4 platforms."""

    def test_render_all_platforms(self):
        from core.response_formatter import (
            Platform, StructuredResponse, ResponseFormatter, TableData,
        )

        fmt = ResponseFormatter()
        r = StructuredResponse(
            title="Pipeline Report",
            body="Showing campaign metrics.",
            status="success",
            metrics={"Leads": 100, "Emails": 50, "Replies": 10},
            tables=[TableData(
                headers=["Campaign", "Leads", "Status"],
                rows=[["Austin Plumbing", "45", "active"]],
                title="Campaigns",
            )],
            progress=0.65,
            agent_name="Reporter",
            agent_emoji="📊",
        )

        # CLI
        cli_out = fmt.render(r, Platform.CLI)
        assert "Pipeline Report" in cli_out
        assert "100" in cli_out
        assert "65%" in cli_out

        # Discord
        dc_out = fmt.render(r, Platform.DISCORD)
        assert "**" in dc_out
        assert "Pipeline Report" in dc_out

        # Telegram
        tg_out = fmt.render(r, Platform.TELEGRAM)
        assert "<b>" in tg_out
        assert "Pipeline Report" in tg_out

        # GUI
        gui_out = fmt.render(r, Platform.GUI)
        assert "<div" in gui_out
        assert "Pipeline Report" in gui_out

    def test_discord_embed(self):
        from core.response_formatter import StructuredResponse, ResponseFormatter

        fmt = ResponseFormatter()
        r = StructuredResponse(
            title="Task Done",
            body="Completed successfully",
            status="success",
            metrics={"Count": 42},
            agent_name="Scout",
            agent_emoji="🔍",
        )
        embed = fmt.render_embed(r)
        assert embed["title"] == "Task Done"
        assert embed["color"] == 0x22C55E
        assert embed["footer"]["text"] == "🔍 Scout"
        assert len(embed["fields"]) == 1


class TestExcelExportIntegration:
    """Test Excel exports with real data."""

    def test_campaign_export_round_trip(self, db):
        from core.excel_export_engine import ExcelExportEngine

        # Seed campaign with leads
        with db.session_scope() as session:
            camp = Campaign(
                name="Integration Export", search_query="plumbing austin",
                total_leads=3, qualified_leads=2, emails_sent=2, replies=1,
                total_token_cost=0.25,
            )
            session.add(camp)
            session.flush()
            for i in range(3):
                session.add(Lead(
                    campaign_id=camp.id, business_name=f"Lead {i}",
                    email=f"lead{i}@test.com", status="qualified" if i < 2 else "new",
                    website_score=80, tier2_cost_usd=0.01,
                ))
            camp_id = camp.id

        engine = ExcelExportEngine(db)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name

        try:
            result = engine.export_campaign_xlsx(camp_id, path)
            assert os.path.exists(result)

            import openpyxl
            wb = openpyxl.load_workbook(result)
            assert "Summary" in wb.sheetnames
            assert "Leads" in wb.sheetnames
            assert wb["Leads"].max_row >= 4  # header + 3 leads
        finally:
            os.unlink(path)

    def test_finance_export_with_invoices(self, db):
        from core.pricing_engine import PricingEngine
        from core.excel_export_engine import ExcelExportEngine

        pe = PricingEngine(db)
        pe.create_invoice(
            client_name="Export A",
            line_items=[{"description": "Svc", "quantity": 1, "unit_price": 500}],
        )
        pe.create_invoice(
            client_name="Export B",
            line_items=[{"description": "Svc", "quantity": 2, "unit_price": 300}],
        )

        engine = ExcelExportEngine(db)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name

        try:
            result = engine.export_finance_xlsx(path)
            assert os.path.exists(result)

            import openpyxl
            wb = openpyxl.load_workbook(result)
            assert "Invoices" in wb.sheetnames
            assert wb["Invoices"].max_row >= 3  # header + 2 invoices
        finally:
            os.unlink(path)


class TestDiscordNotificationIntegration:
    """Test Discord notification routing end-to-end."""

    def test_event_routing(self):
        from core.gateway_adapters.discord_server import DiscordServerManager
        from core.gateway_adapters.discord_notifications import DiscordNotificationRouter

        mgr = DiscordServerManager()
        mgr._guild_configs["test_guild"] = {
            "leads": "ch_leads",
            "outreach": "ch_outreach",
            "invoices": "ch_invoices",
            "logs": "ch_logs",
        }

        adapter = MagicMock()
        router = DiscordNotificationRouter(mgr, adapter)

        # Multiple event types should route to correct channels
        events = [
            ("lead_found", {"name": "Acme"}, "ch_leads"),
            ("email_sent", {"to": "test@test.com"}, "ch_outreach"),
            ("invoice_approved", {"number": "INV-001"}, "ch_invoices"),
        ]

        for event_type, data, expected_channel in events:
            adapter.reset_mock()
            router.route_event(event_type, data)
            call_args = adapter.send_embed.call_args[0]
            assert call_args[0] == expected_channel, (
                f"{event_type} should route to {expected_channel}"
            )


class TestTelegramCommandIntegration:
    """Test Telegram command routing with mocked gateway."""

    def test_command_pipeline(self):
        from core.gateway_adapters.telegram_commands import TelegramCommandRouter

        gw = MagicMock()
        gw.handle_inbound.return_value = "5 campaigns, 100 leads"
        router = TelegramCommandRouter(gateway_engine=gw)

        # Multiple commands
        for cmd in ["stats", "fleet", "invoice"]:
            result = router.handle_command(cmd, "", "owner_123")
            assert result["success"], f"/{cmd} should succeed"

    def test_approval_callback_chain(self):
        from core.gateway_adapters.telegram_commands import TelegramCommandRouter

        gw = MagicMock()
        gw.handle_inbound.return_value = "Action approved"
        router = TelegramCommandRouter(gateway_engine=gw)

        result = router.handle_callback("approve_99", "owner_123")
        assert result["success"]
        assert "approved" in result["response"].lower()


class TestPricingEvaluation:
    """Test pricing evaluation with service catalog."""

    def test_rule_based_with_services(self, db):
        from core.pricing_engine import PricingEngine

        with db.session_scope() as session:
            camp = Campaign(name="Price Test")
            session.add(camp)
            session.flush()
            lead = Lead(
                campaign_id=camp.id, business_name="Eval Corp",
                email="eval@corp.com", website_score=85,
            )
            session.add(lead)
            session.flush()
            lead_id = lead.id

            for name, price in [("Design", 1000), ("SEO", 500), ("Ads", 300)]:
                session.add(Service(
                    name=name, base_price=price, max_price=price * 3,
                    category="marketing",
                ))

        pe = PricingEngine(db)
        result = pe.evaluate_pricing(lead_id)
        assert result["success"]
        assert len(result["data"]["line_items"]) == 3
        assert result["data"]["subtotal"] > 0


class TestCrossFeatureWiring:
    """Test that features work together correctly."""

    def test_invoice_numbers_sequential(self, db):
        from core.pricing_engine import PricingEngine

        pe = PricingEngine(db)

        numbers = []
        for i in range(5):
            r = pe.create_invoice(
                client_name=f"Client {i}",
                line_items=[{"description": "X", "quantity": 1, "unit_price": 100}],
            )
            numbers.append(r["data"]["invoice_number"])

        # Verify sequential numbering
        for i, num in enumerate(numbers):
            assert num == f"INV-{i + 1:04d}"

    def test_notify_event_method_exists(self, db):
        """Verify gateway engine has notify_event method."""
        from core.gateway_engine import GatewayEngine
        from core.key_vault import KeyVault

        ge = GatewayEngine(db, KeyVault())
        assert hasattr(ge, "notify_event")

        # Should not raise even with no adapters
        ge.notify_event("test_event", {"data": "test"})
